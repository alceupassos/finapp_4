#!/usr/bin/env python3
"""
Script para Geração de DRE e DFC para Múltiplas Empresas
=========================================================

Este script processa dados financeiros de múltiplas empresas (CNPJs) e gera
demonstrativos DRE (Demonstrativo de Resultados) e DFC (Demonstrativo de Fluxo de Caixa).

Arquivos necessários:
- PlanoDeContas.xlsx: Plano de contas padrão (compartilhado entre todas empresas)
- CentroDeCustos.xlsx: Centros de custo padrão (compartilhado entre todas empresas)
- [CNPJ].xlsx: Arquivo de dados para cada empresa (formato: 26888098000159.xlsx)

Autor: Sistema Angra Saúde
Data: 2024
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Tuple
import logging

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class ProcessadorDREDFC:
    """
    Classe para processar dados financeiros e gerar DRE e DFC.
    """
    
    def __init__(self, dir_entrada: str, dir_saida: str):
        """
        Inicializa o processador.
        
        Args:
            dir_entrada: Diretório contendo os arquivos de entrada
            dir_saida: Diretório onde serão salvos os arquivos gerados
        """
        self.dir_entrada = Path(dir_entrada)
        self.dir_saida = Path(dir_saida)
        self.dir_saida.mkdir(exist_ok=True, parents=True)
        
        # Carregar arquivos de referência (compartilhados)
        self.plano_contas = None
        self.centro_custos = None
        self.meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
                      'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        
    def carregar_referencias(self):
        """Carrega os arquivos de referência (Plano de Contas e Centro de Custos)."""
        logger.info("Carregando arquivos de referência...")
        
        # Plano de Contas
        plano_path = self.dir_entrada / 'PlanoDeContas.xlsx'
        if not plano_path.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {plano_path}")
        
        self.plano_contas = pd.read_excel(plano_path, header=1)
        logger.info(f"✓ Plano de Contas carregado: {len(self.plano_contas)} contas")
        
        # Centro de Custos
        centro_path = self.dir_entrada / 'CentroDeCustos.xlsx'
        if not centro_path.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {centro_path}")
        
        self.centro_custos = pd.read_excel(centro_path, header=1)
        logger.info(f"✓ Centro de Custos carregado: {len(self.centro_custos)} centros")
        
    def carregar_dados_empresa(self, cnpj: str) -> pd.DataFrame:
        """
        Carrega os dados financeiros de uma empresa específica.
        
        Args:
            cnpj: CNPJ da empresa (sem formatação)
            
        Returns:
            DataFrame com os dados da empresa
        """
        arquivo = self.dir_entrada / f'{cnpj}.xlsx'
        if not arquivo.exists():
            raise FileNotFoundError(f"Arquivo não encontrado: {arquivo}")
        
        # Ler arquivo pulando primeira linha (título geral) e usando segunda como header
        df = pd.read_excel(arquivo, header=1)
        logger.info(f"✓ Dados carregados para CNPJ {cnpj}: {len(df)} registros")
        
        return df
    
    def preparar_dados(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Prepara e limpa os dados para processamento.
        
        Args:
            df: DataFrame com dados brutos
            
        Returns:
            DataFrame preparado
        """
        # Fazer cópia para não modificar original
        df = df.copy()
        
        # Converter datas
        for col in ['Emissão', 'Vencimento', 'Liquidação', 'Competência']:
            if col in df.columns:
                df[col] = pd.to_datetime(df[col], errors='coerce')
        
        # Converter valores numéricos
        for col in ['Valor Bruto', 'Valor Líquido']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')
        
        # Filtrar apenas registros válidos
        df = df[df['Registro'].notna()]
        
        # Adicionar coluna de mês/ano baseado na competência
        if 'Competência' in df.columns:
            df['Mês'] = df['Competência'].dt.month
            df['Ano'] = df['Competência'].dt.year
        
        return df
    
    def agrupar_por_conta_mes(self, df: pd.DataFrame, tipo_demonstrativo: str) -> pd.DataFrame:
        """
        Agrupa dados por conta contábil e mês.
        
        Args:
            df: DataFrame com dados preparados
            tipo_demonstrativo: 'DRE' (regime de competência) ou 'DFC' (regime de caixa)
            
        Returns:
            DataFrame agrupado
        """
        # Definir qual coluna de data usar
        if tipo_demonstrativo == 'DRE':
            # DRE usa competência
            data_col = 'Competência'
        else:  # DFC
            # DFC usa liquidação (caixa)
            data_col = 'Liquidação'
        
        # Filtrar registros com data válida
        df_valido = df[df[data_col].notna()].copy()
        
        # Adicionar mês/ano baseado na data apropriada
        df_valido['Mês'] = df_valido[data_col].dt.month
        df_valido['Ano'] = df_valido[data_col].dt.year
        
        # Agrupar por Plano de Contas e Mês
        agrupado = df_valido.groupby(['Plano de Contas', 'Mês']).agg({
            'Valor Líquido': 'sum'
        }).reset_index()
        
        return agrupado
    
    def criar_estrutura_demonstrativo(self, df_agrupado: pd.DataFrame, 
                                      nome_empresa: str, 
                                      tipo: str) -> pd.DataFrame:
        """
        Cria a estrutura formatada do demonstrativo (DRE ou DFC).
        
        Args:
            df_agrupado: DataFrame com dados agrupados
            nome_empresa: Nome da empresa
            tipo: 'DRE' ou 'DFC'
            
        Returns:
            DataFrame formatado para o demonstrativo
        """
        # Criar tabela pivotada (contas x meses)
        pivot = df_agrupado.pivot_table(
            index='Plano de Contas',
            columns='Mês',
            values='Valor Líquido',
            fill_value=0,
            aggfunc='sum'
        )
        
        # Reordenar colunas para meses 1-12
        colunas_mes = [i for i in range(1, 13) if i in pivot.columns]
        pivot = pivot[colunas_mes]
        
        # Renomear colunas para nomes dos meses
        pivot.columns = [self.meses[i-1] for i in pivot.columns]
        
        # Adicionar coluna Total
        pivot['Total'] = pivot.sum(axis=1)
        
        # Resetar index para ter Plano de Contas como coluna
        pivot = pivot.reset_index()
        
        # TODO: Aqui você pode adicionar lógica para agrupar contas em categorias
        # (Receitas Operacionais, Deduções, Despesas, etc.)
        # Por enquanto, retorna todas as contas
        
        return pivot
    
    def aplicar_formatacao(self, wb: openpyxl.Workbook, ws: openpyxl.worksheet.worksheet.Worksheet,
                          nome_empresa: str, tipo: str):
        """
        Aplica formatação ao demonstrativo no Excel.
        
        Args:
            wb: Workbook do openpyxl
            ws: Worksheet a ser formatada
            nome_empresa: Nome da empresa
            tipo: 'DRE' ou 'DFC'
        """
        # Cores
        cor_cabecalho = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cor_categoria = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # Fonte
        fonte_cabecalho = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        fonte_categoria = Font(name='Calibri', size=11, bold=True)
        fonte_normal = Font(name='Calibri', size=10)
        
        # Alinhamento
        alinhamento_centro = Alignment(horizontal='center', vertical='center')
        alinhamento_esquerda = Alignment(horizontal='left', vertical='center')
        
        # Bordas
        borda_fina = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título da empresa (merge)
        ws.merge_cells('A3:N3')
        cell_titulo = ws['A3']
        cell_titulo.value = f'Nome da Empresa: {nome_empresa}'
        cell_titulo.font = Font(name='Calibri', size=12, bold=True)
        cell_titulo.alignment = alinhamento_esquerda
        
        # Linha em branco
        ws.merge_cells('A4:N4')
        
        # Cabeçalho do demonstrativo (linha 5)
        titulo_demo = 'Demonstrativo de Resultados' if tipo == 'DRE' else 'Demonstrativo de Fluxo de Caixa'
        
        # Aplicar estilo ao cabeçalho (linha 6 - onde estão os meses)
        for col in range(1, 15):  # A até N (14 colunas)
            cell = ws.cell(row=6, column=col)
            cell.fill = cor_cabecalho
            cell.font = fonte_cabecalho
            cell.alignment = alinhamento_centro
            cell.border = borda_fina
        
        # Ajustar largura das colunas
        ws.column_dimensions['A'].width = 50  # Plano de Contas
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
            ws.column_dimensions[col].width = 15
        
        # Formatar valores como moeda
        for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=2, max_col=14):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
                cell.border = borda_fina
    
    def gerar_demonstrativo(self, cnpj: str, nome_empresa: str = None) -> str:
        """
        Gera DRE e DFC para uma empresa específica.
        
        Args:
            cnpj: CNPJ da empresa
            nome_empresa: Nome da empresa (opcional, se não fornecido usa o CNPJ)
            
        Returns:
            Caminho do arquivo gerado
        """
        if nome_empresa is None:
            nome_empresa = f"CNPJ {cnpj}"
        
        logger.info(f"\n{'='*80}")
        logger.info(f"Processando: {nome_empresa} ({cnpj})")
        logger.info(f"{'='*80}")
        
        # Carregar dados da empresa
        df_empresa = self.carregar_dados_empresa(cnpj)
        
        # Preparar dados
        df_preparado = self.preparar_dados(df_empresa)
        
        # Gerar DRE
        logger.info("Gerando DRE (Regime de Competência)...")
        df_dre_agrupado = self.agrupar_por_conta_mes(df_preparado, 'DRE')
        df_dre = self.criar_estrutura_demonstrativo(df_dre_agrupado, nome_empresa, 'DRE')
        
        # Gerar DFC
        logger.info("Gerando DFC (Regime de Caixa)...")
        df_dfc_agrupado = self.agrupar_por_conta_mes(df_preparado, 'DFC')
        df_dfc = self.criar_estrutura_demonstrativo(df_dfc_agrupado, nome_empresa, 'DFC')
        
        # Salvar em Excel
        arquivo_saida = self.dir_saida / f'DRE_DFC_{cnpj}.xlsx'
        logger.info(f"Salvando arquivo: {arquivo_saida}")
        
        with pd.ExcelWriter(arquivo_saida, engine='openpyxl') as writer:
            # Escrever DRE
            df_dre.to_excel(writer, sheet_name='DRE', index=False, startrow=5)
            
            # Escrever DFC
            df_dfc.to_excel(writer, sheet_name='DFC', index=False, startrow=5)
        
        # Aplicar formatação
        logger.info("Aplicando formatação...")
        wb = openpyxl.load_workbook(arquivo_saida)
        
        # Formatar DRE
        ws_dre = wb['DRE']
        self.aplicar_formatacao(wb, ws_dre, nome_empresa, 'DRE')
        
        # Formatar DFC
        ws_dfc = wb['DFC']
        self.aplicar_formatacao(wb, ws_dfc, nome_empresa, 'DFC')
        
        # Salvar
        wb.save(arquivo_saida)
        
        logger.info(f"✓ Demonstrativo gerado com sucesso!")
        logger.info(f"  - DRE: {len(df_dre)} contas")
        logger.info(f"  - DFC: {len(df_dfc)} contas")
        
        return str(arquivo_saida)
    
    def processar_multiplas_empresas(self, lista_cnpjs: List[Tuple[str, str]]):
        """
        Processa múltiplas empresas em lote.
        
        Args:
            lista_cnpjs: Lista de tuplas (cnpj, nome_empresa)
        """
        logger.info(f"\n{'='*80}")
        logger.info(f"PROCESSAMENTO EM LOTE - {len(lista_cnpjs)} empresas")
        logger.info(f"{'='*80}\n")
        
        resultados = []
        erros = []
        
        for cnpj, nome in lista_cnpjs:
            try:
                arquivo = self.gerar_demonstrativo(cnpj, nome)
                resultados.append((cnpj, nome, arquivo))
            except Exception as e:
                logger.error(f"✗ Erro ao processar {nome} ({cnpj}): {str(e)}")
                erros.append((cnpj, nome, str(e)))
        
        # Resumo
        logger.info(f"\n{'='*80}")
        logger.info("RESUMO DO PROCESSAMENTO")
        logger.info(f"{'='*80}")
        logger.info(f"✓ Processados com sucesso: {len(resultados)}/{len(lista_cnpjs)}")
        logger.info(f"✗ Com erros: {len(erros)}/{len(lista_cnpjs)}")
        
        if resultados:
            logger.info("\n✓ Arquivos gerados:")
            for cnpj, nome, arquivo in resultados:
                logger.info(f"  - {nome}: {arquivo}")
        
        if erros:
            logger.info("\n✗ Erros encontrados:")
            for cnpj, nome, erro in erros:
                logger.info(f"  - {nome} ({cnpj}): {erro}")


def main():
    """Função principal para execução do script."""
    
    # Configurações
    DIR_ENTRADA = '/mnt/user-data/uploads'  # Ajustar conforme necessário
    DIR_SAIDA = '/mnt/user-data/outputs'
    
    # Criar processador
    processador = ProcessadorDREDFC(DIR_ENTRADA, DIR_SAIDA)
    
    # Carregar arquivos de referência
    processador.carregar_referencias()
    
    # Exemplo: processar uma empresa
    # processador.gerar_demonstrativo('26888098000159', 'GRUPO VOLPE - MATRIZ')
    
    # Exemplo: processar múltiplas empresas
    # lista_empresas = [
    #     ('26888098000159', 'GRUPO VOLPE - MATRIZ'),
    #     ('CNPJ2', 'EMPRESA 2'),
    #     ('CNPJ3', 'EMPRESA 3'),
    #     # ... adicionar demais empresas
    # ]
    # processador.processar_multiplas_empresas(lista_empresas)
    
    logger.info("\n✓ Script configurado e pronto para uso!")
    logger.info("\nPara processar empresas, utilize:")
    logger.info("  - processador.gerar_demonstrativo(cnpj, nome_empresa)")
    logger.info("  - processador.processar_multiplas_empresas(lista_cnpjs)")


if __name__ == '__main__':
    main()
